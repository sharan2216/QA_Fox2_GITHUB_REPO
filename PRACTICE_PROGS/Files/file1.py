from openpyxl import load_workbook
wb=load_workbook("C:\\Users\\shashi\\Desktop\\ExcelSheet.xlsx")
sh=wb['Login']
print(sh['A1'].value)
print(wb['Login']['A2'].value)
print("--------------------")
print(sh.cell(1,1).value)
print(sh.cell(2,1).value)
print(sh.cell(3,1).value)
print("--------------------")
print(sh.cell(row=1,column=2).value)
print(sh.cell(row=2,column=2).value)
print(sh.cell(row=3,column=2).value)


