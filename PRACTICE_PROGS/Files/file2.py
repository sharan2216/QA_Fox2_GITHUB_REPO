from openpyxl import load_workbook
wb=load_workbook("C:\\Users\\shashi\\Desktop\\ExcelSheet.xlsx")

sh=wb['Login']
row=sh.max_row
col=sh.max_column

print(row)
print(col)

for i in range(1,row+1):
    for j in range(1,col+1):
        print(sh.cell(i,j).value,end=' ')
    print('')
