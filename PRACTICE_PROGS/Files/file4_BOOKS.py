# Pages xpath= //li[@class='current']
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
# from webdriver_manager import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import time
import openpyxl
from selenium.webdriver.support.wait import WebDriverWait

path=r'C:\Users\shashi\Desktop\Bookdata.xlsx'
wb=openpyxl.load_workbook(path)
# sh1=wb.active
sh1=wb['Sheet1']

driver = webdriver.Chrome()
driver.get("https://books.toscrape.com/")
driver.maximize_window()
driver.implicitly_wait(3)
time.sleep(3)

page_title=driver.find_element(By.XPATH,"//li[@class='current']").text
print(page_title)
page_last_num=page_title.split()[-1]
Pages_count=int(page_last_num)
print(Pages_count)
for page in range(1,Pages_count):
    if page ==1:
        Books=driver.find_elements(By.XPATH,"//a[@title]")
        for book in Books:
            book.click()
            driver.implicitly_wait(3)
            book_title = driver.find_element(By.XPATH,"//h1").text
            print(book_title)
            Table_rows = len(driver.find_elements(By.XPATH, "//*[@class='table table-striped']/tbody/tr"))
            print(Table_rows)

            for r in range(1, Table_rows + 1):
                Product_info = driver.find_element(By.XPATH,'//*[@class="table table-striped"]/tbody/tr['+str(r)+']/th').text
                Product_value= driver.find_element(By.XPATH,'//*[@class="table table-striped"]/tbody/tr['+str(r)+']/td').text
                Excel_max_row = sh1.max_row
                sh1.cell(row=Excel_max_row + 1, column=1).value = book_title
                sh1.cell(row=Excel_max_row + 1, column=2).value = Product_info
                sh1.cell(row=Excel_max_row + 1, column=3).value = Product_value
                wb.save(path)
            driver.back()




